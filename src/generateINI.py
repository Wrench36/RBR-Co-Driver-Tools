"""
    Restructured to work with ranges instead of hard coded columns and rows. ini to excel will have to be updated to match.
"""


import pandas as pd
import tkinter as tk
from tkinter import filedialog
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
import os

def get_sheet_names(excel_file):
    xls = pd.ExcelFile(excel_file)
    return xls.sheet_names

def select_sheet(sheet_names, excel_file):
    def on_select(event):
        global selected_sheet
        widget = event.widget
        selection = widget.curselection()
        if selection:
            selected_sheet = widget.get(selection[0])
            process_sheet(excel_file, selected_sheet)
            sheet_names.remove(selected_sheet)
            if not sheet_names:
                root.quit()
            else:
                listbox.delete(0, tk.END)
                for sheet in sheet_names:
                    listbox.insert(tk.END, sheet)

    root = tk.Tk()
    root.title("Select a sheet")

    listbox = tk.Listbox(root, height=len(sheet_names))
    listbox.pack()

    for sheet in sheet_names:
        listbox.insert(tk.END, sheet)

    listbox.bind('<<ListboxSelect>>', on_select)

    root.mainloop()

def defTable_to_DefDict(ws,tableName):
    outputDict = {}
    table = ws.tables[tableName]
    table_range = table.ref
    table_head = ws[table_range][0]
    #print(table_head)
    table_data = ws[table_range][1:]
    columns = [column for column in table_head]
    for cell in columns:
        if cell.value == "name":
            defNameCol = cell.column
    
    for row in table_data:
        noteDict = {} 
        dictKey = row[defNameCol - 1].value
        for cell in row:
            key = table_head[cell.column - 1].value
            val = cell.value
            noteDict[key] = val
        outputDict[dictKey] = noteDict
    return outputDict

def orgTable_to_orgDict(ws,tableName):
    outputDict = {}
    table = ws.tables[tableName]
    table_range = table.ref
    table_head = ws[table_range][0]
    #print(table_head)
    table_data = ws[table_range][1:]
    dictKey = 1
    
    
    for row in table_data:
        noteDict = []
        for cell in row:
            key = len(noteDict)+1
            val = cell.value
            noteDict.append(val)
        outputDict[dictKey] = noteDict
        dictKey = dictKey + 1
    return outputDict

def process_sheet(excel_file, sheet_name):
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]

    defDict = defTable_to_DefDict(ws,f"{sheet_name}Definitions")
    orgDict = orgTable_to_orgDict(ws,f"{sheet_name}Organization")
    
    output_file = f"{sheet_name}.ini"
    current_datetime = datetime.now()
    current_datetime_str = current_datetime.strftime("%Y-%m-%d %H:%M:%S")
    with open(output_file, 'w') as file:
        file.write(f"; {sheet_name}\n; generateINI.py\n; By: Wrench\n; Date: {current_datetime_str}\n\n")
        for key, value in orgDict.items():
            col = 0
            if key != 'Column':
                for note in value:
                    if note != None:
                        print(defDict[note])
                        for entry, value in defDict[note].items():
                            if entry == "name":
                                file.write(f"[PACENOTE::{value}]\n")
                            elif entry != "column" and value != None:
                                 file.write(f"{entry}={value}\n")
                        file.write(f"column={col}\n\n")
                        col = col+1
   

def main():
    root = tk.Tk()
    root.withdraw()
    parent_dir = os.path.dirname(os.path.abspath(__file__)) 
    docs_dir = os.path.abspath(os.path.join(parent_dir, "../"))
    excel_file = filedialog.askopenfilename(initialdir=docs_dir,title="Select Excel File", filetypes=[("Excel files", "*.xlsx")])

    if not excel_file:
        print("No file selected, exiting.")
        return
    
    sheet_names = get_sheet_names(excel_file)

    select_sheet(sheet_names, excel_file)

# Run the main function
if __name__ == "__main__":
    main()

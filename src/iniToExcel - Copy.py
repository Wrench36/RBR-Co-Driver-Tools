import tkinter as tk
from tkinter import filedialog, messagebox
import configparser
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import PatternFill
import os
import subprocess

KEY_ORDER = ["id", "sounds", "snd0", "snd1", "column", "link"]

def select_file_dialog(file_type):
    root = tk.Tk()
    root.withdraw()
    file = filedialog.askopenfilename(filetypes=[(file_type, "*")])
    return file

def select_or_create_excel():
    file_path = select_file_dialog("Excel files")
    if not file_path:
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if not file_path:
        return None
    return file_path

def parse_ini(file_path):
    config = configparser.ConfigParser()
    config.read(file_path)
    data = {}
    for section in config.sections():
        if section.startswith("PACENOTE::"):
            name = section.replace("PACENOTE::", "")
            values = {key: config.get(section, key, fallback="") for key in KEY_ORDER}
            data[name] = values
    return data

def write_to_excel(data, file_path, sheet_name, org_col_start=10):
    workbook = openpyxl.load_workbook(file_path) if os.path.exists(file_path) else openpyxl.Workbook()
    if "Sheet1" in workbook.sheetnames:
        workbook.remove(workbook["Sheet1"])
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])
    sheet = workbook.create_sheet(sheet_name)

    # Set headers for the main data table
    headers = ["Name", "ID", "Sounds", "Snd0", "Snd1", "Column", "Link"]
    for col_num, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_num, value=header)

    # Write pacenotes data to columns A onward, starting from row 2
    row = 2
    for name, values in data.items():
        sheet.cell(row=row, column=1, value=name)  # Section name without "PACENOTE::"
        for col_num, key in enumerate(KEY_ORDER, start=2):
            sheet.cell(row=row, column=col_num, value=values[key])  # Only the value
        row += 1

    # Create header for organization table, merged across five columns starting at column J
    org_start_col_letter = get_column_letter(org_col_start)
    sheet.merge_cells(f"{org_start_col_letter}1:{get_column_letter(org_col_start + 4)}1")
    sheet[f"{org_start_col_letter}1"] = "Organization Table"
    sheet[f"{org_start_col_letter}1"].alignment = Alignment(horizontal="center")

    # Populate the organization table based on 'column' values in each INI section
    org_columns = {}  # Dictionary to keep track of the current row for each column
    for name, values in data.items():
        column_index = int(values.get("column", 0)) + org_col_start  # Column based on 'column' key
        if column_index not in org_columns:
            org_columns[column_index] = 2  # Start from row 2 for each new column
        else:
            org_columns[column_index] += 1  # Move to the next row in the column
        row_index = org_columns[column_index]
        cell = sheet.cell(row=row_index, column=column_index, value=name)
        print(cell)

    
    dv = DataValidation(type="list", formula1="=A2:A100", showDropDown=False)
    dv.add('J2:N10')
    sheet.add_data_validation(dv)

    # Create a fill color for highlighting duplicates
    # Define the green fill color
    highlight_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    highlight_Text = Font(color="006100")
    dxf = DifferentialStyle(font=highlight_Text, fill=highlight_fill)
    # Formula for duplicate checking across ranges $J$2:$N$10 and $A$2:$A$25
    # The COUNTIF formula will check if a value appears more than once across both ranges
    formula_duplicate_check = '=COUNTIF($A$2:$A$100, J2) + COUNTIF($J$2:$N$10, J2) > 1'

    

    # Apply conditional formatting for duplicates in the range $J$2:$N$10 (Across both ranges)
    sheet.conditional_formatting.add(
        "J2:N10", 
        FormulaRule(formula=[formula_duplicate_check], fill=highlight_fill)
    )

    # Apply conditional formatting for duplicates in the range $A$2:$A$25 (Across both ranges)
    sheet.conditional_formatting.add(
        "A2:A25", 
        FormulaRule(formula=[formula_duplicate_check], fill=highlight_fill)
    )


    workbook.save(file_path)



def process_ini_to_excel(excel_file):
    ini_file = select_file_dialog("INI files")
    if not ini_file:
        print("No INI file selected.")
        return False

    data = parse_ini(ini_file)
    sheet_name = os.path.splitext(os.path.basename(ini_file))[0]  # Remove '.ini' extension
    write_to_excel(data, excel_file, sheet_name)
    print(f"Data written to {excel_file} in sheet '{sheet_name}'")
    return True

def main():
    excel_file = select_or_create_excel()
    if not excel_file:
        print("No Excel file selected or created.")
        return

    while True:
        result = process_ini_to_excel(excel_file)
        if not result:
            break

        # Ask if user wants to process another file
        choice = messagebox.askquestion("Continue?", "Would you like to process another INI file?")
        if choice == 'no':
            break

    # Open the Excel file on exit
    subprocess.Popen(['start', excel_file], shell=True)

if __name__ == "__main__":
    main()

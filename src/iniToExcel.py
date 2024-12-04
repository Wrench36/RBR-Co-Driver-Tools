"""
This is close. When the excel file is opened, excel will add "@" to the next id formula and break it.
"""
import tkinter as tk
from tkinter import filedialog, messagebox
import configparser
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Color, PatternFill, Font, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import PatternFill
import os
import re
import subprocess

KEY_ORDER = ["id", "sounds", "snd0", "snd1", "column", "link"]

def select_file_dialog(file_type,initial_dir):
    root = tk.Tk()
    root.withdraw()
    file = filedialog.askopenfilename(initialdir=initial_dir,title=f"Select {file_type}",filetypes=[(file_type, "*")])
    return file

def select_or_create_excel():
    file_path = select_file_dialog("Excel files")
    if not file_path:
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if not file_path:
        return None
    return file_path

def parse_ini(file_path):
    config = configparser.ConfigParser()
    config.read(file_path)
    data = {}
    for section in config.sections():
        if section.startswith("PACENOTE::"):
            name = section.replace("PACENOTE::", "")
            values = {key: config.get(section, key, fallback="") for key in config[section]}
            data[name] = values
    return data

def write_data_table(data,sheet,sheet_name,table_style):
    longest_idx = 0
    longest = 0
    current_idx = 0
    for key, value in data.items():
        current_idx += 1
        itemLen = len(value)

        if itemLen > longest:
            longest = itemLen
            longest_idx = current_idx

    keys = list(data.keys())
    target_key = keys[longest_idx-1]
 
    longestDict = {target_key: data[target_key]}
    header = []
    for key, value in longestDict.items():
        for key2, note in value.items():
            header.append(key2)
    
    for col_num, headers in enumerate(header, start=1):
        sheet.cell(row=1, column=col_num, value=headers)
         
    row = 2

    last_col = 1
    last_col_letter = get_column_letter(last_col)
    for names, values in data.items():
        for name, value in enumerate(values, start=1):
            col_num = header.index(value)+1
            if col_num > last_col:
                last_col = col_num
                last_col_letter = get_column_letter(last_col)
            sheet.cell(row=row, column=col_num, value=values[value])
        row += 1
    
    table_range = f"A1:{last_col_letter}{50}"  # A1 to the last column/row
    table = Table(displayName=f"{sheet_name}Definitions", ref=table_range)
    table.tableStyleInfo = table_style
    sheet.add_table(table)
    return last_col

def write_org_table(data,sheet,sheet_name,last_col,table_style):
    org_start_col_letter = get_column_letter(last_col+2)
    
    org_columns = {}  # Dictionary to keep track of the current row for each column
    for name, values in data.items():
        column_index = int(values.get("column", 0)) + (last_col + 2)
        if column_index not in org_columns:
            org_columns[column_index] = 2
        else:
            org_columns[column_index] += 1
        row_index = org_columns[column_index]
        if isinstance(name, str) and name.isdigit():
            name = int(name)
        sheet.cell(row=row_index, column=column_index, value=name)


    for col_idx in range(last_col+2, last_col+8):
        header_cell = sheet.cell(row=1, column=col_idx)
        if not header_cell.value or not isinstance(header_cell.value, str):
            header_cell.value = f"Column{col_idx - last_col + 3}"
        
    rangeString = f"{org_start_col_letter}{1}:{get_column_letter(last_col+7)}{row_index}"
    table_range = rangeString
    table2 = openpyxl.worksheet.table.Table(ref=table_range,
                                        displayName=f"{sheet_name}Organization",
                                        tableStyleInfo=table_style)
    sheet.add_table(table2)
    
    ##### Data Validation
    dv = DataValidation(type="list", formula1="=$A$2:$A$50", showDropDown=False)
    dv.add(rangeString)
    sheet.add_data_validation(dv)
    
    absolute_range = re.sub(r'([A-Z]+)(\d+)', r'$\1$\2', rangeString)
    return absolute_range

def write_to_excel(data, file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path) if os.path.exists(file_path) else openpyxl.Workbook()
    if "Sheet1" in workbook.sheetnames:
        workbook.remove(workbook["Sheet1"])
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])
    sheet = workbook.create_sheet(sheet_name)
    
    nameDict = {}
    for key, value in data.items():
        nameDict[key] = {}
        nameDict[key].update({"name":key})

        
    for key, value in data.items():
        nameDict[key].update(value)
    data = nameDict
    
    ##### tabe style
    mediumStyle = openpyxl.worksheet.table.TableStyleInfo(name='TableStyleMedium2',
                                                        showRowStripes=True)
    
    last_def_col = write_data_table(data,sheet,sheet_name,mediumStyle)
    org_table_range = write_org_table(data,sheet,sheet_name,last_def_col,mediumStyle)
   
    ##### Conditional Formatting
    highlight_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    formula_duplicate_check = f'=COUNTIF({org_table_range},A2)>0'
    sheet.conditional_formatting.add(
        "A2:A50",
        FormulaRule(formula=[formula_duplicate_check], fill=highlight_fill)
    )
    
    ##### Column width
    for col_num in range(1, last_def_col + 8):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        sheet.column_dimensions[col_letter].auto_size = True
        
        
    #########Next ID Cell
    sheet_names = workbook.sheetnames
    formula = "=MAX(" + ",".join([f"VALUE({sheet}!B2:B100)" for sheet in sheet_names]) + ")+1"
    next_id_col = last_def_col + 1
    next_id_col_ltr = get_column_letter(next_id_col)
    #next_id_cell = sheet[f"{next_id_col_ltr}1"]
    for sheet in workbook.worksheets:
        sheet[f"{next_id_col_ltr}1"].value = "Next ID"
        sheet[f"{next_id_col_ltr}2"].value = formula

    workbook.save(file_path)


def process_ini_to_excel(excel_file):
    parent_dir = os.path.dirname(os.path.abspath(__file__)) 
    packages_dir = os.path.abspath(os.path.join(parent_dir, "../../config/pacenotes/packages"))
    
    ini_file = select_file_dialog("INI files",packages_dir)
    if not ini_file:
        print("No INI file selected.")
        return False

    data = parse_ini(ini_file)
    sheet_name = os.path.splitext(os.path.basename(ini_file))[0]  # Remove '.ini' extension
    write_to_excel(data, excel_file, sheet_name)
    print(f"Data written to {excel_file} in sheet '{sheet_name}'")
    return True

def main():
    parent_dir = os.path.dirname(os.path.abspath(__file__)) 
    docs_dir = os.path.abspath(os.path.join(parent_dir, "../"))
    excel_file = select_file_dialog("Excel files",docs_dir)
    if not excel_file:
        print("No Excel file selected or created.")
        return

    while True:
        result = process_ini_to_excel(excel_file)
        if not result:
            break

if __name__ == "__main__":
    main()
